from formatter import format_excel
from io import BytesIO
from typing import Dict, List

from fastapi import Depends, File, Header, UploadFile
from openpyxl import load_workbook
from pymongo.database import Database

from services.api import OK, Error, responses
from services.dependencies import get_db

from . import api


def load_values_as_array(file: BytesIO) -> List:
    elements = []
    wb = load_workbook(filename=file)
    active_sheet = wb.active
    max_row = active_sheet.max_row
    max_col = active_sheet.max_column

    for r in range(1, max_row + 1):
        element = []
        for c in range(1, max_col + 1):
            element.append(active_sheet.cell(row=r, column=c).value)
        for i in element:
            if i is not None:
                elements.append(element)
                break
    return elements


@api.post('/format_excel', responses=responses)
def format_excel_to_yml(id: int, file: UploadFile = File(...), db: Database = Depends(get_db)):
    elements = load_values_as_array(BytesIO(file.file.read()))

    try:
        instructions = db.find_one({'user_id': str(id)})['instructions']
    except TypeError:
        return Error('Invalid user id')

    return format_excel(elements, instructions)


@api.post('/instructions', responses=responses)
def save_instructions(id: int, instructions: List, db: Database = Depends(get_db)):
    ent = db.find_one({'user_id': str(id)})
    if ent is not None:
        db.update_one(
            {'_id': ent['_id']}, {'$set': {'user_id': str(id), 'instructions': instructions}}
        )
    else:
        db.insert_one({'user_id': str(id), 'instructions': instructions})
    print(instructions)
    return OK(None)


@api.get('/instructions', responses=responses)
def get_instructions(id: int, db: Database = Depends(get_db)):
    try:
        return db.find_one({'user_id': str(id)})['instructions']
    except TypeError:
        return Error('Invalid user id')
